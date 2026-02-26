# cari_app.py
# Tek dosyada: Excel import + SQLite + hesaplama + Tkinter arayüz + PDF ekstre
# Gerekli paketler:
#   pip install openpyxl reportlab
# Çalıştır:
#   python cari_app.py

import sqlite3
from pathlib import Path
from datetime import datetime, date

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas


DB_PATH = Path("cari.db")


# =========================
# DB
# =========================
SCHEMA = """
CREATE TABLE IF NOT EXISTS transactions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    tarih TEXT NOT NULL,               -- YYYY-MM-DD
    musteri TEXT NOT NULL,
    aciklama TEXT,
    islem_turu TEXT NOT NULL,          -- Satış / Alış / Ödeme / Tahsilat
    ayar TEXT,                         -- Has / 925 / 935
    gram REAL DEFAULT 0,
    birim TEXT DEFAULT 'gr',
    iscilik_doviz TEXT,                -- USD / EUR / TL (opsiyon)
    birim_fiyat_veya_nakit REAL DEFAULT 0
);
"""

def db_connect():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con

def db_init():
    with db_connect() as con:
        con.executescript(SCHEMA)

def db_insert_many(rows: list[dict]):
    with db_connect() as con:
        con.executemany(
            """
            INSERT INTO transactions
            (tarih, musteri, aciklama, islem_turu, ayar, gram, birim, iscilik_doviz, birim_fiyat_veya_nakit)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    r["tarih"],
                    r["musteri"],
                    r.get("aciklama", ""),
                    r["islem_turu"],
                    r.get("ayar", ""),
                    float(r.get("gram", 0) or 0),
                    r.get("birim", "gr"),
                    (r.get("iscilik_doviz", None) or None),
                    float(r.get("birim_fiyat_veya_nakit", 0) or 0),
                )
                for r in rows
            ],
        )

def db_list_customers() -> list[str]:
    with db_connect() as con:
        cur = con.execute("SELECT DISTINCT musteri FROM transactions ORDER BY musteri")
        return [r["musteri"] for r in cur.fetchall()]

def db_get_transactions(musteri: str) -> list[dict]:
    with db_connect() as con:
        cur = con.execute(
            "SELECT * FROM transactions WHERE musteri=? ORDER BY tarih, id",
            (musteri,),
        )
        return [dict(r) for r in cur.fetchall()]

def db_add_transaction(r: dict):
    db_insert_many([r])

def db_delete_transaction(tx_id: int):
    with db_connect() as con:
        con.execute("DELETE FROM transactions WHERE id=?", (tx_id,))


# =========================
# Excel IO
# =========================
def _to_iso_date(v) -> str:
    if v is None or v == "":
        return date.today().isoformat()
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            pass
    return date.today().isoformat()

def import_from_excel(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)

    if "CARİ HAREKETLER" not in wb.sheetnames:
        raise ValueError("Excel içinde 'CARİ HAREKETLER' sayfası bulunamadı.")

    ws = wb["CARİ HAREKETLER"]

    # Beklenen kolonlar (sende bu şekildeydi):
    # A:Tarih B:Müşteri C:Açıklama D:İşlem Türü E:Ayar F:Gram G:Birim H:İşçilik Döviz I:İşçilik Birim Fiyat / Alınan Nakit
    rows = []
    for r in range(2, ws.max_row + 1):
        musteri = ws.cell(r, 2).value
        if not musteri:
            continue
        rows.append(
            {
                "tarih": _to_iso_date(ws.cell(r, 1).value),
                "musteri": str(musteri).strip(),
                "aciklama": str(ws.cell(r, 3).value or "").strip(),
                "islem_turu": str(ws.cell(r, 4).value or "").strip(),
                "ayar": str(ws.cell(r, 5).value or "").strip(),
                "gram": ws.cell(r, 6).value or 0,
                "birim": str(ws.cell(r, 7).value or "gr").strip(),
                "iscilik_doviz": str(ws.cell(r, 8).value or "").strip(),
                "birim_fiyat_veya_nakit": ws.cell(r, 9).value or 0,
            }
        )
    return rows


# =========================
# Hesaplama (Excel mantığı)
# =========================
def ayar_katsayi(ayar: str) -> float:
    if not ayar:
        return 0.0
    a = str(ayar).strip().lower()
    if a == "has":
        return 1.0
    if a in ("925", "0.925"):
        return 0.925
    if a in ("935", "0.935"):
        return 0.935
    return 0.0

def sign_has_gram(islem_turu: str) -> int:
    # Satış/Ödeme + ; Alış/Tahsilat -
    t = (islem_turu or "").strip().lower()
    return 1 if t in ("satış", "satis", "ödeme", "odeme") else -1

def sign_iscilik_tutar(islem_turu: str) -> int:
    # Satış: + ; Alış: - ; Ödeme: + ; Tahsilat: -
    t = (islem_turu or "").strip().lower()
    if t in ("satış", "satis"):
        return 1
    if t in ("alış", "alis"):
        return -1
    if t in ("ödeme", "odeme"):
        return 1
    return -1

def compute_running(transactions: list[dict]) -> list[dict]:
    bakiye_has = 0.0
    bakiye_usd = 0.0
    bakiye_eur = 0.0
    bakiye_tl = 0.0

    out = []
    for tx in transactions:
        islem_turu = tx.get("islem_turu", "")
        gram = float(tx.get("gram", 0) or 0)
        ayar = tx.get("ayar", "")
        doviz = (tx.get("iscilik_doviz") or "").strip().upper()
        i = float(tx.get("birim_fiyat_veya_nakit", 0) or 0)

        has = gram * ayar_katsayi(ayar) * sign_has_gram(islem_turu)
        bakiye_has += has

        # Satış/Alış: gram * i ; Ödeme/Tahsilat: i
        if (islem_turu or "").strip().lower() in ("satış", "satis", "alış", "alis"):
            iscilik_tutar = sign_iscilik_tutar(islem_turu) * gram * i
        else:
            iscilik_tutar = sign_iscilik_tutar(islem_turu) * i

        if doviz == "USD":
            bakiye_usd += iscilik_tutar
        elif doviz == "EUR":
            bakiye_eur += iscilik_tutar
        elif doviz == "TL":
            bakiye_tl += iscilik_tutar

        row = dict(tx)
        row["iscilik_tutar"] = iscilik_tutar
        row["has_gram"] = has
        row["bakiye_has"] = bakiye_has
        row["bakiye_usd"] = bakiye_usd
        row["bakiye_eur"] = bakiye_eur
        row["bakiye_tl"] = bakiye_tl
        out.append(row)

    return out


# =========================
# PDF
# =========================
def export_statement_pdf(pdf_path: str, musteri: str, computed_rows: list[dict]):
    c = canvas.Canvas(pdf_path, pagesize=A4)
    w, h = A4

    y = h - 50
    c.setFont("Helvetica-Bold", 14)
    c.drawString(40, y, f"Cari Ekstre - {musteri}")
    y -= 30

    if computed_rows:
        last = computed_rows[-1]
        c.setFont("Helvetica", 10)
        c.drawString(40, y, f"Has Gram Bakiye: {last['bakiye_has']:.3f} gr")
        y -= 14
        c.drawString(40, y, f"USD: {last['bakiye_usd']:.2f}   EUR: {last['bakiye_eur']:.2f}   TL: {last['bakiye_tl']:.2f}")
        y -= 22

    c.setFont("Helvetica-Bold", 9)
    c.drawString(40, y, "Tarih")
    c.drawString(100, y, "İşlem")
    c.drawString(170, y, "Ayar")
    c.drawString(220, y, "Gram")
    c.drawString(280, y, "Döviz")
    c.drawString(330, y, "Tutar")
    c.drawString(400, y, "Has Bakiye")
    y -= 12
    c.line(40, y, w - 40, y)
    y -= 14

    c.setFont("Helvetica", 8)
    for tx in computed_rows[-55:]:
        if y < 60:
            c.showPage()
            y = h - 50
            c.setFont("Helvetica", 8)

        c.drawString(40, y, str(tx.get("tarih", "")))
        c.drawString(100, y, str(tx.get("islem_turu", ""))[:10])
        c.drawString(170, y, str(tx.get("ayar", ""))[:6])
        c.drawRightString(260, y, f"{float(tx.get('gram', 0) or 0):.3f}")
        c.drawString(280, y, str(tx.get("iscilik_doviz", ""))[:3])
        c.drawRightString(380, y, f"{float(tx.get('iscilik_tutar', 0) or 0):.2f}")
        c.drawRightString(500, y, f"{float(tx.get('bakiye_has', 0) or 0):.3f}")
        y -= 12

    c.save()


# =========================
# UI (Tkinter)
# =========================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Cari Takip (Excel -> Uygulama)")
        self.geometry("1100x650")

        db_init()

        top = ttk.Frame(self)
        top.pack(fill="x", padx=10, pady=8)

        ttk.Button(top, text="Excel İçe Aktar", command=self.import_excel).pack(side="left")
        ttk.Button(top, text="PDF Ekstre", command=self.export_pdf).pack(side="left", padx=6)

        ttk.Label(top, text="Müşteri:").pack(side="left", padx=(20, 6))
        self.customer = ttk.Combobox(top, values=db_list_customers(), width=35, state="readonly")
        self.customer.pack(side="left")
        self.customer.bind("<<ComboboxSelected>>", lambda e: self.refresh())

        ttk.Button(top, text="Yeni Kayıt", command=self.add_dialog).pack(side="right")
        ttk.Button(top, text="Seçili Kaydı Sil", command=self.delete_selected).pack(side="right", padx=6)

        summary = ttk.LabelFrame(self, text="Son Durum")
        summary.pack(fill="x", padx=10, pady=(0, 8))
        self.lbl_summary = ttk.Label(summary, text="—")
        self.lbl_summary.pack(anchor="w", padx=10, pady=6)

        cols = ("id","tarih","aciklama","islem_turu","ayar","gram","doviz","birim_fiyat","iscilik_tutar","has_bakiye","usd","eur","tl")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=22)
        self.tree.pack(fill="both", expand=True, padx=10, pady=10)

        headings = {
            "id":"ID","tarih":"Tarih","aciklama":"Açıklama","islem_turu":"İşlem Türü","ayar":"Ayar",
            "gram":"Gram","doviz":"Döviz","birim_fiyat":"Birim Fiyat/Nakit","iscilik_tutar":"İşçilik Tutar",
            "has_bakiye":"Has Bakiye","usd":"USD B.","eur":"EUR B.","tl":"TL B."
        }
        for c in cols:
            self.tree.heading(c, text=headings.get(c, c))
            self.tree.column(c, width=95 if c != "aciklama" else 280, anchor="w")

        self.tree.tag_configure("mavi", background="#d7eaff")     # Satış/Ödeme
        self.tree.tag_configure("kirmizi", background="#ffd6d6")  # Tahsilat/Alış

        vals = db_list_customers()
        if vals:
            self.customer.set(vals[0])
            self.refresh()

    def import_excel(self):
        path = filedialog.askopenfilename(title="Excel seç", filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        try:
            rows = import_from_excel(path)
            if not rows:
                messagebox.showwarning("Uyarı", "Excel'de import edilecek kayıt bulunamadı.")
                return
            db_insert_many(rows)
            self.customer["values"] = db_list_customers()
            if not self.customer.get().strip():
                self.customer.set(db_list_customers()[0])
            self.refresh()
            messagebox.showinfo("Tamam", f"{len(rows)} kayıt içe aktarıldı.")
        except Exception as e:
            messagebox.showerror("Hata", str(e))

    def refresh(self):
        musteri = self.customer.get().strip()
        if not musteri:
            return

        tx_rows = db_get_transactions(musteri)
        computed = compute_running(tx_rows)

        for item in self.tree.get_children():
            self.tree.delete(item)

        for tx in computed:
            islem = (tx.get("islem_turu") or "").lower().strip()
            tag = ""
            if islem in ("tahsilat", "alış", "alis"):
                tag = "kirmizi"
            if islem in ("satış", "satis", "ödeme", "odeme"):
                tag = "mavi"

            self.tree.insert(
                "", "end",
                values=(
                    tx["id"],
                    tx.get("tarih",""),
                    tx.get("aciklama",""),
                    tx.get("islem_turu",""),
                    tx.get("ayar",""),
                    float(tx.get("gram",0) or 0),
                    (tx.get("iscilik_doviz") or ""),
                    float(tx.get("birim_fiyat_veya_nakit",0) or 0),
                    float(tx.get("iscilik_tutar",0) or 0),
                    float(tx.get("bakiye_has",0) or 0),
                    float(tx.get("bakiye_usd",0) or 0),
                    float(tx.get("bakiye_eur",0) or 0),
                    float(tx.get("bakiye_tl",0) or 0),
                ),
                tags=(tag,)
            )

        if computed:
            last = computed[-1]
            self.lbl_summary.config(
                text=(
                    f"Has Gram: {last['bakiye_has']:.3f} gr | "
                    f"USD: {last['bakiye_usd']:.2f} | EUR: {last['bakiye_eur']:.2f} | TL: {last['bakiye_tl']:.2f} | "
                    f"Son İşlem: {last.get('tarih','')} / {last.get('islem_turu','')}"
                )
            )
        else:
            self.lbl_summary.config(text="Kayıt yok.")

    def add_dialog(self):
        musteri = self.customer.get().strip()
        if not musteri:
            messagebox.showwarning("Uyarı", "Önce müşteri seçin.")
            return

        win = tk.Toplevel(self)
        win.title("Yeni Kayıt")
        win.geometry("450x380")

        def row(label, r):
            ttk.Label(win, text=label).grid(row=r, column=0, sticky="w", padx=10, pady=6)

        row("Tarih (YYYY-MM-DD)", 0)
        e_tarih = ttk.Entry(win)
        e_tarih.insert(0, date.today().isoformat())
        e_tarih.grid(row=0, column=1, padx=10, pady=6)

        row("Açıklama", 1)
        e_acik = ttk.Entry(win)
        e_acik.grid(row=1, column=1, padx=10, pady=6)

        row("İşlem Türü", 2)
        cb_islem = ttk.Combobox(win, values=["Satış","Alış","Ödeme","Tahsilat"], state="readonly")
        cb_islem.set("Satış")
        cb_islem.grid(row=2, column=1, padx=10, pady=6)

        row("Ayar", 3)
        cb_ayar = ttk.Combobox(win, values=["Has","925","935"], state="readonly")
        cb_ayar.set("Has")
        cb_ayar.grid(row=3, column=1, padx=10, pady=6)

        row("Gram", 4)
        e_gram = ttk.Entry(win)
        e_gram.insert(0, "0")
        e_gram.grid(row=4, column=1, padx=10, pady=6)

        row("İşçilik Döviz (USD/EUR/TL)", 5)
        e_doviz = ttk.Entry(win)
        e_doviz.grid(row=5, column=1, padx=10, pady=6)

        row("Birim Fiyat / Nakit", 6)
        e_i = ttk.Entry(win)
        e_i.insert(0, "0")
        e_i.grid(row=6, column=1, padx=10, pady=6)

        def save():
            try:
                # basit kontrol
                _ = datetime.strptime(e_tarih.get().strip(), "%Y-%m-%d")
                gram = float(e_gram.get().strip().replace(",", "."))
                tutar = float(e_i.get().strip().replace(",", "."))

                db_add_transaction({
                    "tarih": e_tarih.get().strip(),
                    "musteri": musteri,
                    "aciklama": e_acik.get().strip(),
                    "islem_turu": cb_islem.get().strip(),
                    "ayar": cb_ayar.get().strip(),
                    "gram": gram,
                    "birim": "gr",
                    "iscilik_doviz": e_doviz.get().strip(),
                    "birim_fiyat_veya_nakit": tutar,
                })
                win.destroy()
                self.refresh()
            except Exception as e:
                messagebox.showerror("Hata", str(e))

        ttk.Button(win, text="Kaydet", command=save).grid(row=8, column=1, sticky="e", padx=10, pady=14)

    def delete_selected(self):
        sel = self.tree.selection()
        if not sel:
            return
        tx_id = int(self.tree.item(sel[0])["values"][0])
        if messagebox.askyesno("Sil", f"ID {tx_id} kaydı silinsin mi?"):
            db_delete_transaction(tx_id)
            self.refresh()

    def export_pdf(self):
        musteri = self.customer.get().strip()
        if not musteri:
            return
        computed = compute_running(db_get_transactions(musteri))

        path = filedialog.asksaveasfilename(
            title="PDF kaydet",
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")]
        )
        if not path:
            return
        export_statement_pdf(path, musteri, computed)
        messagebox.showinfo("Tamam", "PDF oluşturuldu.")


if __name__ == "__main__":
    App().mainloop()
